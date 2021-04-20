import csv
import os
import config
import tool
import logging
import time
from openpyxl import Workbook


def read_csv(csv_path):
    csvFile = open(csv_path, 'r')
    reader = csv.reader(csvFile)

    result = []
    # 倒序循环
    for item in reader:
        if reader.line_num == 1:
            continue
        result.append({'Date': item[0], 'Adj Close': item[5]})

    csvFile.close()
    return reversed(result)


def new_excel():
    source_path = 'Source'
    date_type = 'mo'
    g = os.walk(source_path)
    # path,dir_list,file_list

    for i, val in enumerate(g):
        if i == 0:
            # 获取类别列表
            dirCategoryList = val[1]

    workbook = Workbook()
    # 删除第一张默认生成的 sheet
    workbook.remove(workbook['Sheet'])
    # 类别循环
    for category_name in dirCategoryList:
        # 新建 sheet
        sheet = workbook.create_sheet()
        sheet.title = category_name
        category_path = '%s\\%s\\1%s' % (source_path, category_name, date_type)
        category_path_g = os.walk(category_path)
        column_index = 1
        # 基金循环
        for path, dir_list, file_list in category_path_g:
            for csv_file in file_list:
                fund_name = csv_file.split('-')[0]
                sheet.cell(row=1, column=column_index+1).value = fund_name
                sheet.cell(row=1, column=column_index +
                           2).value = '%s return' % fund_name
                # 构建单个基金 cvs 文件路径
                fund_cvs_path = '%s\\%s' % (category_path, csv_file)
                row_index_1 = 0
                row_index_2 = 0
                for item in read_csv(fund_cvs_path):
                    row_index_1 = row_index_1 + 1
                    sheet.cell(row=row_index_1 + 1,
                               column=1).value = item['Date']
                    if item['Adj Close'] == 'null':
                        continue
                    sheet.cell(row=row_index_1 + 1, column=column_index +
                               1).value = float(item['Adj Close'])
                for item in read_csv(fund_cvs_path):
                    row_index_2 = row_index_2 + 1
                    if sheet.cell(row=row_index_2 + 2, column=column_index+1).value == None:
                        continue
                    yesterday_adj_close = float(sheet.cell(
                        row=row_index_2 + 2, column=column_index+1).value)
                    if item['Adj Close'] == 'null':
                        continue
                    sheet.cell(row=row_index_2 + 1, column=column_index +
                               2).value = float((float(item['Adj Close']) - yesterday_adj_close) / yesterday_adj_close)
                column_index = column_index + 2

    # 保存为 xlsx
    workbook.save('D:\\Code\\Python\\fund_crawler\\finally-%s-%s.xlsx' %
                  (date_type, tool.get_now_time()))


logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s \tFile \'%(filename)s\'[line:%(lineno)d] %(levelname)s %(message)s',
    datefmt='%a, %d %b %Y %H:%M:%S',
    filename='./' + time.strftime('%Y-%m-%d',
                                  time.localtime(time.time())) + '.log',
    filemode='a')


new_excel()

# source_path = 'Source'

# g = os.walk(source_path)
# # path,dir_list,file_list

# for i, val in enumerate(g):
#     if i == 0:
#         dirCategoryList = val[1]
#     print(val[2])
